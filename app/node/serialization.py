# File: app/nodes/serialization.py
from dotenv import load_dotenv
load_dotenv()
import os
import datetime
from xml.etree.ElementTree import Element, SubElement, indent, tostring
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from app.state import GraphState
try:
    from langsmith import traceable
except ImportError:
    def traceable(name=None):
        def decorator(fn):
            return fn
        return decorator


def build_xml(state: GraphState) -> str:
    """
    Builds a clinical XML document conforming to a simplified CDISC-inspired schema.
    In a real deployment this would reference the formal E2B R3 or CDISC CDASH schema.
    """

    root = Element("ClinicalDocument")
    root.set("xmlns", "urn:medistream:clinical:v1")
    root.set("created", datetime.datetime.utcnow().isoformat())

    # Medications section
    meds_section = SubElement(root, "ConcomitantMedications")
    for med in state.enriched_medications:
        med_el = SubElement(meds_section, "Medication")
        SubElement(med_el, "GenericName").text = med.generic_name
        SubElement(med_el, "ActiveComposition").text = med.active_composition
        SubElement(med_el, "Dose").text = med.original.dose or ""
        SubElement(med_el, "Route").text = med.original.route or ""
        SubElement(med_el, "Frequency").text = med.original.frequency or ""
        SubElement(med_el, "StartDate").text = med.original.start_date.isoformat() if med.original.start_date else ""
        SubElement(med_el, "EndDate").text = med.original.end_date.isoformat() if med.original.end_date else ""
        SubElement(med_el, "ATCCode").text = med.atc_code or ""
        SubElement(med_el, "IsConcomitant").text = str(med.original.is_concomitant)

    # Medical history section
    history_section = SubElement(root, "MedicalHistory")
    for event in state.enriched_history:
        ev_el = SubElement(history_section, "HistoryEvent")
        SubElement(ev_el, "Description").text = event.original.event_description
        SubElement(ev_el, "MedDRA_LLT").text = event.meddra_llt
        SubElement(ev_el, "MedDRA_PT").text = event.meddra_pt
        SubElement(ev_el, "MedDRA_SOC").text = event.meddra_soc or ""
        SubElement(ev_el, "ConfidenceScore").text = str(round(event.confidence_score, 4))
        SubElement(ev_el, "OnsetDate").text = event.original.onset_date.isoformat() if event.original.onset_date else ""
        SubElement(ev_el, "IsOngoing").text = str(event.original.is_ongoing)

    # Timeline section
    timeline_section = SubElement(root, "ChronologicalTimeline")
    for entry in state.chronological_timeline:
        entry_el = SubElement(timeline_section, "TimelineEntry")
        SubElement(entry_el, "EventType").text = entry.event_type
        SubElement(entry_el, "Description").text = entry.description
        SubElement(entry_el, "ISOTimestamp").text = entry.iso_timestamp

    indent(root, space="  ")
    return tostring(root, encoding="unicode", xml_declaration=True)


def build_excel(state: GraphState) -> openpyxl.Workbook:
    """
    Creates a formatted Excel workbook with two sheets:
    one for medications and one for medical history.
    """

    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    # Sheet 1: Medications
    ws_meds = wb.active
    ws_meds.title = "Concomitant Medications"
    med_headers = [
        "Generic Name", "Active Composition", "ATC Code",
        "Dose", "Route", "Frequency", "Start Date", "End Date", "Concomitant"
    ]

    for col, header in enumerate(med_headers, start=1):
        cell = ws_meds.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for row_idx, med in enumerate(state.enriched_medications, start=2):
        ws_meds.cell(row=row_idx, column=1, value=med.generic_name)
        ws_meds.cell(row=row_idx, column=2, value=med.active_composition)
        ws_meds.cell(row=row_idx, column=3, value=med.atc_code or "")
        ws_meds.cell(row=row_idx, column=4, value=med.original.dose or "")
        ws_meds.cell(row=row_idx, column=5, value=med.original.route or "")
        ws_meds.cell(row=row_idx, column=6, value=med.original.frequency or "")
        ws_meds.cell(row=row_idx, column=7, value=med.original.start_date.isoformat() if med.original.start_date else "")
        ws_meds.cell(row=row_idx, column=8, value=med.original.end_date.isoformat() if med.original.end_date else "")
        ws_meds.cell(row=row_idx, column=9, value="Yes" if med.original.is_concomitant else "No")

    for col in ws_meds.columns:
        ws_meds.column_dimensions[col[0].column_letter].width = 22

    # Sheet 2: Medical History
    ws_hist = wb.create_sheet(title="Medical History")
    hist_headers = [
        "Event Description", "MedDRA LLT", "MedDRA PT", "MedDRA SOC",
        "Confidence Score", "Onset Date", "Ongoing"
    ]

    for col, header in enumerate(hist_headers, start=1):
        cell = ws_hist.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for row_idx, event in enumerate(state.enriched_history, start=2):
        ws_hist.cell(row=row_idx, column=1, value=event.original.event_description)
        ws_hist.cell(row=row_idx, column=2, value=event.meddra_llt)
        ws_hist.cell(row=row_idx, column=3, value=event.meddra_pt)
        ws_hist.cell(row=row_idx, column=4, value=event.meddra_soc or "")
        ws_hist.cell(row=row_idx, column=5, value=round(event.confidence_score, 4))
        ws_hist.cell(row=row_idx, column=6, value=event.original.onset_date.isoformat() if event.original.onset_date else "")
        ws_hist.cell(row=row_idx, column=7, value="Yes" if event.original.is_ongoing else "No")

    for col in ws_hist.columns:
        ws_hist.column_dimensions[col[0].column_letter].width = 26

    return wb


@traceable(name="serialization_node")
def serialization_node(state: GraphState) -> GraphState:
    """
    Node 4: Serialization
    Writes XML and Excel files to the outputs directory and updates state with paths.
    """

    os.makedirs("outputs", exist_ok=True)
    timestamp = datetime.datetime.utcnow().strftime("%Y%m%d_%H%M%S")

    xml_path = f"outputs/medistream_{timestamp}.xml"
    xlsx_path = f"outputs/medistream_{timestamp}.xlsx"

    xml_content = build_xml(state)
    with open(xml_path, "w", encoding="utf-8") as f:
        f.write(xml_content)

    wb = build_excel(state)
    wb.save(xlsx_path)

    return state.model_copy(update={
        "output_xml_path": xml_path,
        "output_xlsx_path": xlsx_path
    })